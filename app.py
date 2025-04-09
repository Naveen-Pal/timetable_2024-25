from flask import Flask, render_template, request, send_file, session, jsonify
import pandas as pd
from PIL import Image, ImageDraw, ImageFont
import os
import uuid
from datetime import datetime
import logging
import traceback
import io
import csv
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    handlers=[logging.StreamHandler()]
)
logger = logging.getLogger(__name__)

app = Flask(__name__, template_folder='templates')
app.secret_key = os.urandom(24)

output_dir = 'generated_images'
os.makedirs(output_dir, exist_ok=True)

try:
    time_slots = pd.read_csv('Time Slots.csv')
    timetable_data = pd.read_csv('Updated_Processed_Timetable.csv')    
    logger.info("Timetable data loaded successfully")
except Exception as e:
    logger.error(f"Error loading data: {e}")
    timetable_data = pd.DataFrame()
    time_slots = pd.DataFrame()

@app.route('/api/courses', methods=['GET'])
def get_courses():
    try:
        courses = [
            {
                'code': row['Course Code'], 
                'name': row['Course Name'], 
                'credits': row['Credit'],
                'lecture': str(row.get('Lecture Time', '')),
                'tutorial': str(row.get('Tutorial Time', '')),
                'lab': str(row.get('Lab Time', '')),
                'lecture_location': str(row.get('Lecture Location', '')),
                'tutorial_location': str(row.get('Tutorial Location', '')),
                'lab_location': str(row.get('Lab Location', ''))
            } 
            for _, row in timetable_data.iterrows() 
            if not pd.isna(row['Credit'])
        ]
        
        time_slots_csv = pd.read_csv('Time Slots.csv')
        
        days = list(time_slots.columns)
        slots = []
        for idx, row in time_slots.iterrows():
            slot_data = {}
            for day in days:
                slot_data[day] = row[day]
            slots.append(slot_data)
            
        time_slots_column = time_slots_csv.iloc[:, 0].tolist()
        
        return jsonify({
            "courses": courses,
            "days": days,
            "slots": slots,
            "timeLabels": time_slots_column
        })
        
    except Exception as e:
        logger.error(f"Error in get_courses: {e}")
        logger.error(traceback.format_exc())
        return jsonify({"error": str(e)}), 500

@app.route('/', methods=['GET'])
def index():
    try:
        session_id = session.get('session_id')
        if not session_id:
            session_id = str(uuid.uuid4())
            session['session_id'] = session_id
            
        return render_template('index.html', session_id=session_id)
    except Exception as e:
        logger.error(f"Error in index route: {e}")
        logger.error(traceback.format_exc())
        return render_template('error.html', error="Failed to load timetable data")

@app.route('/api/timetable/csv', methods=['POST'])
def generate_timetable_csv():
    try:
        selected_courses = request.form.getlist('courses')
        if not selected_courses:
            return jsonify({"error": "No courses selected"}), 400
            
        try:
            # Create CSV in memory
            csv_data = io.StringIO()
            personalized_timetable = create_timetable_data(selected_courses)
            
            writer = csv.writer(csv_data)
            # Write header - no "Unnamed: 0" column
            writer.writerow(['Time Slot'] + list(personalized_timetable.columns))
            
            # Add time slots as the first column
            time_slots_csv = pd.read_csv('Time Slots.csv')
            time_slots_column = time_slots_csv.iloc[:, 0].tolist()
            
            # Write data rows
            for idx, row in personalized_timetable.iterrows():
                if idx < len(time_slots_column):  # Make sure we don't go out of bounds
                    writer.writerow([time_slots_column[idx]] + list(row))
                
            csv_data.seek(0)
            return send_file(
                io.BytesIO(csv_data.getvalue().encode('utf-8')),
                mimetype='text/csv',
                as_attachment=True,
                download_name=f"Timetable_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
            )
            
        except Exception as e:
            logger.error(f"Error creating CSV timetable: {e}")
            logger.error(traceback.format_exc())
            return jsonify({"error": "Failed to generate CSV timetable"}), 500
            
    except Exception as e:
        logger.error(f"Error in generate_timetable_csv: {e}")
        logger.error(traceback.format_exc())
        return jsonify({"error": str(e)}), 500

@app.route('/api/timetable/excel', methods=['POST'])
def generate_timetable_excel():
    try:
        selected_courses = request.form.getlist('courses')
        if not selected_courses:
            return jsonify({"error": "No courses selected"}), 400
            
        try:
            personalized_timetable = create_timetable_data(selected_courses)
            
            workbook = openpyxl.Workbook()
            worksheet = workbook.active
            worksheet.title = "Timetable"
            
            worksheet.cell(row=1, column=1, value="Time Slot")
            
            for col, day in enumerate(personalized_timetable.columns, 2):
                cell = worksheet.cell(row=1, column=col, value=day)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="0066CC", end_color="0066CC", fill_type="solid")
                cell.font = Font(color="FFFFFF", bold=True)
                cell.alignment = Alignment(horizontal='center')
            
            time_slots_csv = pd.read_csv('Time Slots.csv')
            time_slots_column = time_slots_csv.iloc[:, 0].tolist()
            
            for row_idx, time_slot in enumerate(time_slots_column, 2):
                worksheet.cell(row=row_idx, column=1, value=time_slot)
            
            for row_idx, (_, row) in enumerate(personalized_timetable.iterrows(), 2):
                for col_idx, value in enumerate(row, 2):
                    cell = worksheet.cell(row=row_idx, column=col_idx, value=value)
                    if "(Clash)" in str(value):
                        cell.fill = PatternFill(start_color="FFD6D6", end_color="FFD6D6", fill_type="solid")
                    cell.alignment = Alignment(wrap_text=True)
            
            for col in worksheet.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    if cell.value:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                adjusted_width = (max_length + 2)
                worksheet.column_dimensions[column].width = adjusted_width
            
            excel_data = io.BytesIO()
            workbook.save(excel_data)
            excel_data.seek(0)
            
            return send_file(
                excel_data,
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                as_attachment=True,
                download_name=f"Timetable_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            )
            
        except Exception as e:
            logger.error(f"Error creating Excel timetable: {e}")
            logger.error(traceback.format_exc())
            return jsonify({"error": "Failed to generate Excel timetable"}), 500
            
    except Exception as e:
        logger.error(f"Error in generate_timetable_excel: {e}")
        logger.error(traceback.format_exc())
        return jsonify({"error": str(e)}), 500

@app.route('/api/timetable/text', methods=['POST'])
def get_timetable_text():
    try:
        selected_courses = request.json.get('courses')
        if not selected_courses or not isinstance(selected_courses, list):
            return jsonify({"error": "No courses selected or invalid format"}), 400
            
        try:
            personalized_timetable = create_timetable_data(selected_courses)
            
            time_slots_csv = pd.read_csv('Time Slots.csv')
            time_slots_column = time_slots_csv.iloc[:, 0].tolist()
            
            text_data = []
            
            # Include Time Slot and the actual day columns (no redundant "Time Slot")
            header = ["Time Slot"] + list(personalized_timetable.columns)
            text_data.append("\t".join(header))
            
            for idx, row in personalized_timetable.iterrows():
                # Ensure time_slots_column aligns with rows and is not added twice
                if idx < len(time_slots_column):
                    text_row = [time_slots_column[idx]] + [str(cell).replace('\n', ' ') for cell in row]
                    text_data.append("\t".join(text_row))
            
            return jsonify({"text": "\n".join(text_data)})
            
        except Exception as e:
            logger.error(f"Error creating text timetable: {e}")
            logger.error(traceback.format_exc())
            return jsonify({"error": "Failed to generate text timetable"}), 500
            
    except Exception as e:
        logger.error(f"Error in get_timetable_text: {e}")
        logger.error(traceback.format_exc())
        return jsonify({"error": str(e)}), 500

def create_timetable_data(selected_courses):
    """Common function to create the timetable data structure"""
    course_slots = {}
    for _, row in timetable_data.iterrows():
        if row['Course Code'] in selected_courses:
            course_slots[row['Course Code']] = {
                'name': row['Course Name'],
                'Lecture': str(row.get('Lecture Time', '')),
                'Tutorial': str(row.get('Tutorial Time', '')),
                'Lab': str(row.get('Lab Time', '')),
                'Lecture_Location': str(row.get('Lecture Location', '')),
                'Tutorial_Location': str(row.get('Tutorial Location', '')),
                'Lab_Location': str(row.get('Lab Location', ''))
            }

    # Create a copy of the time slots dataframe without the index column
    personalized_timetable = time_slots.copy()
    # Ensure there's no index column in the output
    personalized_timetable.reset_index(drop=True, inplace=True)
    overlaps = {}

    for slot in personalized_timetable.index:
        for day in personalized_timetable.columns:
            entries = []
            for code, info in course_slots.items():
                for session_type in ['Lecture', 'Tutorial', 'Lab']:
                    times = info.get(session_type)
                    if pd.isna(times) or times == 'nan':
                        continue
                        
                    if personalized_timetable.at[slot, day] in [time.strip() for time in times.split(',') if time.strip()]:
                        location = info.get(f"{session_type}_Location", "")
                        location_text = f"\n{location}" if location and location != "nan" else ""
                        entries.append(f"{code}\n{info['name']}\n{session_type}{location_text}")

            if len(entries) > 1:
                overlaps[f"{slot} {day}"] = entries
                display = ""
                for ent in entries:
                    # Show full course code in clashes too
                    display += ent.split("\n")[0].strip() + "/ "
                
                display = display[:-2] + "\n(Clash)"
                personalized_timetable.at[slot, day] = display
            elif entries:
                personalized_timetable.at[slot, day] = entries[0]
    
    return personalized_timetable

def create_timetable_image(selected_courses, session_id):
    personalized_timetable = create_timetable_data(selected_courses)
    
    output_path = os.path.join(output_dir, f'Timetable_{session_id}.png')
    generate_image(personalized_timetable, output_path)
    return output_path

def generate_image(dataframe, output_path):
    cell_width = 180
    cell_height = 80
    margin = 4
    font_size = 14
    small_font_size = 12
    header_color = (0, 102, 204)
    cell_color = (245, 245, 245)
    clash_color = (255, 220, 220)
    text_color = (0, 0, 0)
    border_color = (200, 200, 200)

    cols = len(dataframe.columns)
    rows = len(dataframe) + 1
    
    image_width = cols * (cell_width + margin) + margin
    image_height = rows * (cell_height + margin) + margin
    
    image = Image.new("RGB", (image_width, image_height), "white")
    draw = ImageDraw.Draw(image)

    try:
        font = ImageFont.truetype("Arial", font_size)
        small_font = ImageFont.truetype("Arial", small_font_size)
    except IOError:
        try:
            font = ImageFont.truetype("/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf", font_size)
            small_font = ImageFont.truetype("/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf", small_font_size)
        except IOError:
            font = ImageFont.load_default()
            small_font = ImageFont.load_default()

    try:
        bold_font = ImageFont.truetype("Arial Bold", font_size + 2)
    except IOError:
        try:
            bold_font = ImageFont.truetype("/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf", font_size + 2)
        except IOError:
            bold_font = font

    # Draw day headers with both full name and abbreviation
    for i, column in enumerate(dataframe.columns):
        x = margin + i * (cell_width + margin)
        y = margin
        draw.rectangle([x, y, x + cell_width, y + cell_height], fill=header_color)
        
        # Full day name
        full_text = column
        abbr_text = column[0] if len(column) > 0 else column
        
        text_width = draw.textlength(full_text, font=bold_font)
        text_x = x + (cell_width - text_width) / 2
        text_y = y + (cell_height - font_size) / 2
        
        draw.text((text_x, text_y), full_text, fill="white", font=bold_font)

    dataframe.reset_index(drop=True, inplace=True)

    # Get time slots for left column
    time_slots_csv = pd.read_csv('Time Slots.csv')
    time_slots_column = time_slots_csv.iloc[:, 0].tolist()

    # Draw time slots on left
    x_timeslot = margin
    for idx, time_slot in enumerate(time_slots_column):
        if idx < len(dataframe):
            y = margin + (idx + 1) * (cell_height + margin)
            draw.rectangle([0, y, margin + cell_width//2, y + cell_height], 
                          fill=cell_color, outline=border_color)
            draw.text((5, y + 5), time_slot, fill=text_color, font=small_font)

    # Draw the actual timetable cells
    for index, row in dataframe.iterrows():
        y = margin + (index + 1) * (cell_height + margin)
        
        for i, column in enumerate(dataframe.columns):
            cell_text = str(row[column])
            x = margin + i * (cell_width + margin)
            
            if "(Clash)" in cell_text:
                current_cell_color = clash_color
            else:
                current_cell_color = cell_color
                
            draw.rectangle([x, y, x + cell_width, y + cell_height], 
                          fill=current_cell_color, outline=border_color)
            
            lines = cell_text.split('\n')
            line_height = font_size + 2
            start_y = y + 5
            
            for j, line in enumerate(lines):
                draw.text((x + 5, start_y + j * line_height), line, fill=text_color, font=font)

    image.save(output_path)

@app.errorhandler(404)
def page_not_found(e):
    return jsonify({"error": "Not found"}), 404

@app.errorhandler(500)
def server_error(e):
    return jsonify({"error": "Internal server error"}), 500

if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5000, debug=False)
